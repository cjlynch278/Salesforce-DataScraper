'''
Created on Feb 26, 2018
@author: 310290474
'''

from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import datetime
import selenium.common.exceptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

envWb = openpyxl.load_workbook('eCC environments Support TEST.xlsx')
wb= openpyxl.load_workbook('Salesforce org disk space TEST.xlsx')
driver = webdriver.Firefox(executable_path='geckodriver.exe') 

now= datetime.datetime.now().date()
prettyDate = str(now.month)+"/"+str(now.day)+"/"+str(now.year)
sheets = iter(wb.worksheets)
sheetsCountIter = iter(wb.worksheets)
sheetCount =  sum(1 for x in sheetsCountIter)

#Skip INstructions
next(sheets)

#InfoRow is the row pulled from "eC environents support" to gather info like the password, username, etc.
infoRow = 0
#cancelled determines if the programs needs to quit or not
cancelled = False
#msRun determines if the master scheduler is running 
msRun = False

#The function that goes through the list of scheduled jobs and determines if the MS is running
def searchMS(msRun):
    try:
        element = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="bodyCell"]/div[4]/div/div[2]/table/tbody/tr[1]/th[6]/a/img')))
        element.click()
           
    except selenium.common.exceptions.TimeoutException:
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/div[2]/table/tbody/tr/td[2]/div[5]/div/div[2]/table/tbody/tr[1]/th[6]/a')))
            element.click()
        except:
            print("Error Sorting Master Scheduler")
            
    i = 2
    element= wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="bodyCell"]/div[4]/div/div[2]/table/tbody/tr['+str(i)+']/th')))           
    
    try:
        #find a way to get count more cleanly
        while( element != None):          
            if 'MasterScheduler scheduled:' in element.text: 
                    element= driver.find_element_by_xpath('//*[@id="bodyCell"]/div[4]/div/div[2]/table/tbody/tr['+str(i)+']/td[5]')
                    if str(prettyDate) in str(element.text):
                        workingRow.append("RUNNING")
                        msRun = True
                        break
            i = i+1
            element= driver.find_element_by_xpath('//*[@id="bodyCell"]/div[4]/div/div[2]/table/tbody/tr['+str(i)+']/th')
   
    except NoSuchElementException:
       emptyVariable = 0     
            
    return msRun

#Exits the program and saves the workbooks           
def exitScrape(cancelled):
    try: 
        if cancelled == False:
            wb.save('Salesforce org disk space TEST.xlsx') 
            print("Work Saved")
        else:   
            print("Workbook not saved")  
    except PermissionError:
        input("Permission denied to save workbook. Workbook may be open. Close it and press 'Enter' to continue")
        #find a way to retry saving 
        wb.save('Salesforce org disk space TEST.xlsx')
        print("Work Saved")
        #  print("Permission Denied to save workbook. Workbook may be open. Please Close and try again")
    wb.close()
    envWb.close()
    exit()
    

#Iterates through all of the sheets in "Salesforce org disk spaces"
for sheet in sheets:
    #Get the first available sheet
    keyValue = sheet.cell(row = 2, column =17).value
    while keyValue == None:
        print("Skipping " + sheet.title)
        try:
            sheet = next(sheets)
            keyValue = sheet.cell(row = 2, column =17).value
        except StopIteration:
            print("End of Sheets reached")
            exitScrape(cancelled)
    
    envSheet =  envWb['Summary List of eCC Env']    
    workingSheet = sheet
    workingRowNum = sum(1 for x in workingSheet.rows)+1
    workingRow = [prettyDate]
    infoRow = 0
    #Iterates through "eCC environments Support" rows and finds a key that matches the keyValue
    while infoRow == 0:
        for i in range(1, sum(1 for x in envSheet.rows)):
            print(str(keyValue) + " + " + str(envSheet.cell(row = i, column = 45).value))
            if str(keyValue) == str(envSheet.cell(row = i, column = 45).value) and keyValue is not None: #DOn't think the keyValue is not None statement
                infoRow = i
                break
        
        if infoRow == 0:       
            print(sheet.title)
            sheet = next(sheets)
            keyValue = sheet.cell(row = 2, column =17).value
    
    #extract useful info from the infoRow
    orgCell = envSheet.cell(row = infoRow, column =1)
    usrCell = envSheet.cell(row = infoRow, column =6)
    pwdCell = envSheet.cell(row = infoRow, column =7)
    pwd = pwdCell.value
    org = orgCell.value
    usr = usrCell.value
    canceled = False
    
    #Scrape Salesforce using the credentials
    print("Checking: " + org + " and writing to: " + sheet.title)
    driver.get("https://login.salesforce.com/")
    try:
        elem = driver.find_element_by_id("username")
        elem.send_keys(usr)
        elem = driver.find_element_by_id("password")
        elem.send_keys(pwd)
        elem = driver.find_element_by_id("Login").click()
        
        wait = WebDriverWait(driver, 60)
        if ".my.salesforce.com/_ui/identity/verification/" in driver.current_url:
            vCode = input("Please enter Verification code (Whitelist your URL in the future): ")
            try:
                elem = driver.find_element_by_id("emc")
                elem.send_keys(vCode)
                driver.find_element_by_id("save").click()
                    
            except selenium.common.exceptions.NoSuchElementException:
                print("Verification Code element not found")    
        element = wait.until(EC.element_to_be_clickable((By.ID, 'setupLink')))
        element.click()
    #Excepts any error that may happen during login
    except:
        #Bad coding practice? Previous code: selenium.common.exceptions.TimeoutException, selenium.common.exceptions.ElementNotInteractableException:
        input("Error occured when logging in. Navigate to the eCC admin homescreen and press 'Enter' here when finished")
        element = wait.until(EC.element_to_be_clickable((By.ID, 'setupLink')))
        element.click()
    
    element = wait.until(EC.element_to_be_clickable((By.ID, 'DataManagement_icon')))
    element.click()
    
    element = wait.until(EC.element_to_be_clickable((By.ID, 'CompanyResourceDisk_font')))
    element.click()
    
    for i in range(2, 4):
        for x in range(1,4):
            element = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="bodyCell"]/div[3]/div[1]/div/div[2]/table/tbody/tr['+str(i)+']/td['+str(x)+']')))
            workingRow.append(element.text)
       
         
    element = wait.until(EC.element_to_be_clickable((By.ID, 'CompanyProfile_icon')))
    element.click()
    element = wait.until(EC.element_to_be_clickable((By.ID, 'CompanyProfileInfo_font')))
    element.click()
    i = 2
    element = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/div[2]/table/tbody/tr/td[2]/div[6]/div[1]/div/div[2]/table/tbody/tr[' +str(i)+']/th')))
    print("Crucial Element: " + element.text)
    #orgID = str(envSheet.cell(row = infoRow, column = 10).value)
    orgID = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/div[2]/table/tbody/tr/td[2]/div[5]/div[2]/div[2]/table/tbody/tr[12]/td[4]')))
    orgID = orgID.text
    retry = 0
    while(int(retry) < 2):    
        try:
            while element is not None:
                print(element.text)
                
                if 'Salesforce' in element.text or 'Customer' in element.text:
                    print("Element text = " + element.text)
                    for x in range(2,5):
                       # print("Element text = " + element.text)
                       # print('//*[@id="'+ orgID+ '_RelatedUserLicenseList_body"]/table/tbody/tr['+str(i)+']/td['+str(x)+']')
                        element = driver.find_element_by_xpath('//*[@id="'+ orgID+ '_RelatedUserLicenseList_body"]/table/tbody/tr['+str(i)+']/td['+str(x)+']')
                            ##C/html/body/div/div[2]/table/tbody/tr/td[2]/div[6]/div[1]/div/div[2]/table/tbody/tr[2]/td[2]
                            ## //*[@id="00D410000007qRa_RelatedUserLicenseList_body"]/table/tbody/tr[2]/td[2]
                        workingRow.append(element.text)
                        element = driver.find_element_by_xpath('//*[@id="'+orgID+ '_RelatedUserLicenseList_body"]/table/tbody/tr[' +str(i)+']/th')
                i = i+1
                element = driver.find_element_by_xpath('//*[@id="'+ orgID+ '_RelatedUserLicenseList_body"]/table/tbody/tr[' +str(i)+']/th')
                print("Next Element " + element.text)
            break        
        except selenium.common.exceptions.NoSuchElementException:
            if(workingRow.__len__() <= 7):
                print("Could not find all elements, check if this is correct SF org ID: " + orgID)
                orgID = input("Please Re-enter OrgID: ")
                retry = retry +1
            else:
                break
       
            
    workingRow.append('')
    element = wait.until(EC.element_to_be_clickable((By.ID, 'Jobs_icon')))
    element.click()
    element = wait.until(EC.element_to_be_clickable((By.ID, 'ScheduledJobs_font')))
    element.click()
    
    msRun = False
    msRun = searchMS(msRun)
    msRun = searchMS(msRun)
    

    if not msRun:
        print("NOTICE Master Scheduler Not Running!")
        workingRow.append("NOT Running")
    
    print("Data Extracted: " )
    print(workingRow)
    workingSheet.append(workingRow)
    
    infoRow = infoRow+1
    print("")
    if envSheet.cell(row = infoRow, column = 1).value is None: 
        print("Reached end of user list")
        break

exitScrape(cancelled)                           